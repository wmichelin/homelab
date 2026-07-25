#!/usr/bin/env python3
"""Normalize Movie Library.xlsx and import into Radarr (Ultra-HD, monitored).

Does NOT search/download on add — use Radarr → Wanted → Missing later in batches.
"""
from __future__ import annotations

import csv
import json
import os
import re
import sys
import time
import urllib.error
import urllib.parse
import urllib.request
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path.home() / "code" / "homelab"
OUT_DIR = ROOT / "backups" / "movie-import"
XLSX = Path.home() / "Downloads" / "Movie Library.xlsx"
RADARR = os.environ.get("RADARR_URL", "http://127.0.0.1:7878")
QUALITY_PROFILE_ID = int(os.environ.get("RADARR_QUALITY_PROFILE_ID", "5"))  # Ultra-HD
ROOT_FOLDER = os.environ.get("RADARR_ROOT_FOLDER", "/movies")
SEARCH_ON_ADD = os.environ.get("RADARR_SEARCH_ON_ADD", "0") == "1"
SLEEP = float(os.environ.get("RADARR_LOOKUP_SLEEP", "0.15"))

YEAR_IN_TITLE = re.compile(r"\s*\((\d{4})\)\s*$")
EDITION_SUFFIX = re.compile(
    r"\s*\((?:Unrated|Uncut(?:\s+Version)?|Extended(?:\s+Cut)?|Director'?s\s+Cut|"
    r"Theatrical(?:\s+Cut)?|Final\s+Cut|Ultimate(?:\s+Cut)?|Special\s+Edition|"
    r"Remastered|IMAX|4K|UHD)\)\s*$",
    re.I,
)
# Sheet title → preferred TMDB id when lookup is ambiguous/broken
TITLE_TMDB_HINTS = {
    "alien 3": 8077,
    "alien³": 8077,
}

def api_key() -> str:
    cfg = ROOT / "media-stack" / "config" / "radarr" / "config.xml"
    m = re.search(r"<ApiKey>([^<]+)</ApiKey>", cfg.read_text())
    if not m:
        raise SystemExit(f"No ApiKey in {cfg}")
    return m.group(1)


def req(method: str, path: str, body=None, key: str = ""):
    data = None if body is None else json.dumps(body).encode()
    url = RADARR + path
    r = urllib.request.Request(
        url,
        data=data,
        method=method,
        headers={"X-Api-Key": key, "Content-Type": "application/json"},
    )
    try:
        with urllib.request.urlopen(r, timeout=60) as resp:
            raw = resp.read().decode()
            return json.loads(raw) if raw else None
    except urllib.error.HTTPError as e:
        err = e.read().decode()[:800]
        raise RuntimeError(f"{method} {path} -> {e.code}: {err}") from e


def normalize_title(raw) -> tuple[str, int | None]:
    """Return (clean_title, year_or_None)."""
    if raw is None:
        return "", None
    if isinstance(raw, float):
        # Excel often stores 1917 as 1917.0
        if raw.is_integer():
            raw = str(int(raw))
        else:
            raw = str(raw)
    elif isinstance(raw, int):
        raw = str(raw)
    else:
        raw = str(raw).strip()

    year = None
    m = YEAR_IN_TITLE.search(raw)
    if m:
        year = int(m.group(1))
        raw = YEAR_IN_TITLE.sub("", raw).strip()

    # Do NOT treat a numeric title (1917, 2012, 1984) as the release year —
    # those are movie titles. Year only comes from "(YYYY)" suffix.

    # Strip edition tags so "American Pie (Unrated)" → "American Pie"
    raw = EDITION_SUFFIX.sub("", raw).strip()

    # Common cleanup
    raw = re.sub(r"\s+", " ", raw).strip()
    # Normalize fancy apostrophes/quotes
    raw = raw.replace("’", "'").replace("‘", "'").replace("“", '"').replace("”", '"')
    # Alien 3 → Alien³ for better TMDB hits when year known later
    if raw.lower() in ("alien 3", "alien3"):
        raw = "Alien³"
    return raw, year


def load_rows(path: Path) -> list[dict]:
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    header = [str(h).strip() if h is not None else "" for h in rows[0]]
    out = []
    for row in rows[1:]:
        if not any(c is not None and str(c).strip() for c in row):
            continue
        d = {header[i]: row[i] if i < len(row) else None for i in range(len(header)) if header[i]}
        title, year = normalize_title(d.get("Title"))
        if not title:
            continue
        owned_raw = d.get("Owned")
        owned = str(owned_raw).strip().lower() if owned_raw is not None else ""
        out.append(
            {
                "title_raw": d.get("Title"),
                "title": title,
                "year": year,
                "director": (str(d.get("Director") or "").strip() or None),
                "collection": (str(d.get("Collection") or "").strip() or None),
                "genre": (str(d.get("Genre") or "").strip() or None),
                "owned": owned in ("yes", "y", "true", "1"),
                "owned_raw": owned_raw,
            }
        )
    return out


def pick_lookup(results: list, title: str, year: int | None, director: str | None):
    if not results:
        return None
    # Hard TMDB hints
    hint = TITLE_TMDB_HINTS.get(title.lower())
    if hint:
        for r in results:
            if r.get("tmdbId") == hint:
                return r, 1000

    title_l = title.lower()
    scored = []
    for idx, r in enumerate(results):
        t = (r.get("title") or "").lower()
        y = r.get("year") or 0
        pop = float(r.get("popularity") or 0)
        score = 0.0
        if t == title_l:
            score += 100
        elif title_l in t or t in title_l:
            score += 40
        if year and y == year:
            score += 50
        elif year and y and abs(y - year) <= 1:
            score += 20
        # Prefer Radarr's popularity ranking; preserve list order as weak signal
        score += min(pop, 50) * 0.5
        score += max(0, 10 - idx) * 0.1
        scored.append((score, pop, -idx, r))
    scored.sort(key=lambda x: (x[0], x[1], x[2]), reverse=True)
    best_score, _pop, _idx, best = scored[0]
    # Reject terrible matches (e.g. Alien 3 → random doc)
    if best_score < 40:
        return None
    return best, best_score


def main() -> int:
    if not XLSX.exists():
        print(f"Missing {XLSX}", file=sys.stderr)
        return 1

    OUT_DIR.mkdir(parents=True, exist_ok=True)
    key = api_key()

    # Resolve root folder id
    roots = req("GET", "/api/v3/rootfolder", key=key) or []
    root = next((r for r in roots if r.get("path").rstrip("/") == ROOT_FOLDER.rstrip("/")), None)
    if not root:
        raise SystemExit(f"Root folder {ROOT_FOLDER} not found: {roots}")

    rows = load_rows(XLSX)
    # Prefer owned; still include unknown-owned
    print(f"Normalized {len(rows)} rows from {XLSX.name}")

    norm_csv = OUT_DIR / "movie-library-normalized.csv"
    with norm_csv.open("w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(
            f,
            fieldnames=["title", "year", "director", "genre", "collection", "owned", "title_raw"],
        )
        w.writeheader()
        for r in rows:
            w.writerow(
                {
                    "title": r["title"],
                    "year": r["year"] or "",
                    "director": r["director"] or "",
                    "genre": r["genre"] or "",
                    "collection": r["collection"] or "",
                    "owned": "yes" if r["owned"] else "",
                    "title_raw": r["title_raw"],
                }
            )
    print(f"Wrote {norm_csv}")

    existing = req("GET", "/api/v3/movie", key=key) or []
    by_tmdb = {m["tmdbId"]: m for m in existing if m.get("tmdbId")}
    by_title_year = {(m.get("title", "").lower(), m.get("year")): m for m in existing}

    added = skipped = unmatched = errors = 0
    unmatched_rows = []
    results_log = []

    for i, row in enumerate(rows, 1):
        title, year = row["title"], row["year"]
        # Build search term
        if title.lower() in ("alien³", "alien 3", "alien3"):
            q = "Alien 3 1992"
        elif year:
            q = f"{title} {year}"
        else:
            q = title
        try:
            lookups = req(
                "GET",
                "/api/v3/movie/lookup?term=" + urllib.parse.quote(q),
                key=key,
            ) or []
        except Exception as e:
            errors += 1
            unmatched_rows.append({**row, "reason": f"lookup error: {e}"})
            print(f"[{i}/{len(rows)}] LOOKUP FAIL {title}: {e}")
            time.sleep(SLEEP)
            continue

        picked = pick_lookup(lookups, title, year, row["director"])
        if not picked:
            unmatched += 1
            unmatched_rows.append({**row, "reason": "no lookup results"})
            print(f"[{i}/{len(rows)}] UNMATCHED {title}")
            time.sleep(SLEEP)
            continue

        movie, score = picked
        tmdb = movie.get("tmdbId")
        if not tmdb:
            unmatched += 1
            unmatched_rows.append({**row, "reason": "no tmdbId"})
            print(f"[{i}/{len(rows)}] UNMATCHED(no tmdb) {title}")
            time.sleep(SLEEP)
            continue

        if tmdb in by_tmdb or (title.lower(), movie.get("year")) in by_title_year:
            skipped += 1
            results_log.append(
                {
                    "status": "exists",
                    "sheet_title": title,
                    "sheet_year": year,
                    "matched": movie.get("title"),
                    "matched_year": movie.get("year"),
                    "tmdbId": tmdb,
                    "score": score,
                }
            )
            if i % 50 == 0:
                print(f"[{i}/{len(rows)}] skip existing…")
            time.sleep(SLEEP * 0.5)
            continue

        # Radarr expects the lookup payload plus library fields
        payload = dict(movie)
        payload.update(
            {
                "qualityProfileId": QUALITY_PROFILE_ID,
                "rootFolderPath": ROOT_FOLDER,
                "monitored": True,
                "minimumAvailability": "released",
                "addOptions": {
                    "searchForMovie": SEARCH_ON_ADD,
                    "monitor": "movieOnly",
                },
            }
        )

        try:
            added_movie = req("POST", "/api/v3/movie", body=payload, key=key)
            added += 1
            by_tmdb[tmdb] = added_movie
            results_log.append(
                {
                    "status": "added",
                    "sheet_title": title,
                    "sheet_year": year,
                    "matched": movie.get("title"),
                    "matched_year": movie.get("year"),
                    "tmdbId": tmdb,
                    "score": score,
                }
            )
            print(
                f"[{i}/{len(rows)}] ADD {movie.get('title')} ({movie.get('year')}) "
                f"← {title!r} score={score}"
            )
        except Exception as e:
            errors += 1
            unmatched_rows.append({**row, "reason": f"add error: {e}"})
            print(f"[{i}/{len(rows)}] ADD FAIL {title}: {e}")

        time.sleep(SLEEP)

    unmatched_csv = OUT_DIR / "movie-library-unmatched.csv"
    with unmatched_csv.open("w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(
            f,
            fieldnames=["title", "year", "director", "genre", "owned", "reason", "title_raw"],
        )
        w.writeheader()
        for r in unmatched_rows:
            w.writerow(
                {
                    "title": r.get("title"),
                    "year": r.get("year") or "",
                    "director": r.get("director") or "",
                    "genre": r.get("genre") or "",
                    "owned": "yes" if r.get("owned") else "",
                    "reason": r.get("reason"),
                    "title_raw": r.get("title_raw"),
                }
            )

    log_json = OUT_DIR / "movie-library-import-log.json"
    log_json.write_text(json.dumps(results_log, indent=2))

    print("\n=== SUMMARY ===")
    print(f"sheet rows:   {len(rows)}")
    print(f"added:        {added}")
    print(f"already had:  {skipped}")
    print(f"unmatched:    {unmatched}")
    print(f"errors:       {errors}")
    print(f"normalized:   {norm_csv}")
    print(f"unmatched:    {unmatched_csv}")
    print(f"log:          {log_json}")
    print("Search/download NOT triggered. In Radarr: Wanted → Missing → search in batches.")
    return 0 if errors == 0 else 2


if __name__ == "__main__":
    raise SystemExit(main())
